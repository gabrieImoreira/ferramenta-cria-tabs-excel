from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import re, traceback, time
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.filters import AutoFilter, FilterColumn
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font

# variáveis
aba_principal = 'aba principal'
tabela_principal = 'TabelaBase'
cabecalho_linha = '19'
range_config_tabelas = 'A7:D12'
range_colunas_para_criar = 'A7:A12'
range_posicao_colunas_para_criar = 'F7:BC12'
range_formatacao_colunas = 'F6:BC6'
range_largura_colunas = 'F7:BC7'
range_total_row = 'F15:BC15'
largura_padrao_colunas = 3

# variáveis importantes
cabecalho = []
linhas_tabela_principal = []
linhas_config_tabelas = []
linhas_posicao_colunas_para_criar = []
linha_formatacao_colunas = []
linha_largura_colunas = []

linha_total_row = []

def remove_table_filters(table: Table, ws: Worksheet) -> None:
    # Only remove filters if there is a header
    if not table.headerRowCount:
        return

    # Initialize table if not done (from WorksheetWriter.write_tables)
    if not table.tableColumns:
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        for idx in range(min_col, max_col + 1):
            col = TableColumn(id=idx, name=f"Column{idx}")
            table.tableColumns.append(col)
        table.autoFilter = AutoFilter(ref=table.ref)
        try:
            row = ws[table.ref][0]
            for cell, col in zip(row, table.tableColumns):
                if cell.data_type != "s":
                    warn("File may not be readable: column headings must be strings.")
                col.name = str(cell.value)
        except TypeError:
            warn("Column headings are missing, file may not be readable")

    filter_columns = table.autoFilter.filterColumn
    # clear any current FilterColumns that might be there
    filter_columns.clear()
    for idx in range(len(table.tableColumns)):
        # ColId of filters is relative to Table, so starts at 0
        # Remove filter with hiddenButton=True
        filter_columns.append(FilterColumn(colId=idx, hiddenButton=True, filters=None))

def transform_cell(cell, tab_name):
    if "SUBTOTAL" in cell:
        # =SUBTOTAL(106,S1E1_tabelaY[FR])
        # =SUBTOTAL(109;[FR])
        list_values = cell.split(',[')
        new_value = f'{list_values[0]},{tab_name}[{list_values[1]}'
        return new_value[1:]
    else:
        start_value = cell

        list_values = cell.split('[@')
        formated_values = []
        for index, value in enumerate(list_values):

            if ']' in value:
                value = value.split(']')[0]
                old_value = f"[@{value}]"
                new_value = f"({tab_name}[[#This Row],[{value}]])"
                formated_values.append([old_value, new_value])

        for value in formated_values:
            start_value = start_value.replace(value[0], value[1])
        return start_value[1:]

if __name__ == "__main__":
    # Ler o arquivo de entrada
    try:
        print('Lendo arquivo de entrada...')
        with open('params.txt', 'r') as file:
            for line in file:
                if '=' in line:
                    key, value = line.strip().split('=')
                    if key.strip() == 'ARQUIVO_ENTRADA':
                        arquivo_entrada = value.strip()
                    elif key.strip() == 'ARQUIVO_SAIDA':
                        arquivo_saida = value.strip()

        if not arquivo_entrada.endswith('.xlsx'):
            raise ValueError('O arquivo de entrada deve ser um arquivo Excel (.xlsx)')
        # Abrindo a primeira aba pra coleta de informações
        wb = load_workbook(arquivo_entrada, data_only=True)
        ws = wb[aba_principal]
        # print(ws.tables.items())
        table_range = ws.tables[tabela_principal]

        print('Lendo informações da planilha...')
        
        # Pegar cabeçalho
        range_cabecalho = table_range.ref.replace('20', cabecalho_linha)
        for row in ws[range_cabecalho]:
            for cell in row:
                cabecalho.append(cell.value)
            break
        
        # Iterar sobre as linhas da tabela principal
        for linha_excel in ws[table_range.ref]:
            linha = []
            for cell in linha_excel:
                linha.append(cell.value)
            linhas_tabela_principal.append(linha)

        # Obter tabela de configuração
        for linha_excel in ws[range_config_tabelas]:
            linha = []
            for cell in linha_excel:
                linha.append(cell.value)
            linhas_config_tabelas.append(linha)

        # Obter tabela de posição de configuração
        for linha_excel in ws[range_posicao_colunas_para_criar]:
            linha = []
            for cell in linha_excel:
                linha.append(cell.value)
            linhas_posicao_colunas_para_criar.append(linha)
        
        # Obter linhas de formatação de colunas
        for linha_excel in ws[range_formatacao_colunas]:
            for cell in linha_excel:
                linha_formatacao_colunas.append(cell.value)

        # Obter linhas de formatação de colunas
        for linha_excel in ws[range_largura_colunas]:
            for cell in linha_excel:
                linha_largura_colunas.append(cell.value)

        # Obter linha de total row
        for linha_excel in ws[range_total_row]:
            for cell in linha_excel:
                linha_total_row.append(cell.value)

        print('Criando arquivo de saída...')
        # Começar a escrever no arquivo de saída
        tabelas_principais = []
        linha_tabela = []
        for index, linha in enumerate(linhas_tabela_principal):
            if linha[1] == 'x':
                if index != 0:
                    tabelas_principais.append(linha_tabela)
                linha_tabela = []
                linha_tabela.append(linha)
            else:
                linha_tabela.append(linha)
        tabelas_principais.append(linha_tabela)
        
                
        del tabelas_principais[0]
        index_filhos_tabela = -1

        print('Criando tabelas...')
        for index_principal, linha in enumerate(linhas_tabela_principal):
            index_filhos_tabela = -1
            if linha[1] == 'x' and index_filhos_tabela <= index_principal:
                print('Criando tabela...', linha[4])
                # print('Linha:', linha)
                # Criar nova aba
                nova_aba = wb.create_sheet(title=linha[4])
                # Ajustar tamanho das colunas
                for i in range(1, 200):
                    nova_aba.column_dimensions[get_column_letter(i)].width = largura_padrao_colunas
                # inserir valores padrão
                nova_aba['U10'] = linha[4]
                nova_aba['U11'] = linha[5]
                nova_aba['U12'] = linha[6]
                nova_aba['U13'] = linha[7]
                nova_aba['U14'] = linha[8]

                # Criar primeira tabela
                # print('index_filhos_tabela', index_filhos_tabela)
                for linha_config, tabela in enumerate(linhas_config_tabelas):
                    index_filhos_tabela = index_principal + 1
                    if linha_config > 0 and tabela[2] is not None:
                        # print(linha_config, tabela[2], tabela[3])
                        # print('aqui')
                        # tabela[2] # nome da tabela
                        # tabela[3] # posição da tabela
                        # posicao_tabela = tabela[3]
                        coluna_inicial = re.findall(r'[a-zA-Z]', tabela[3])
                        coluna_inicial = column_index_from_string(''.join(coluna_inicial))
                        linha_inicial = int(''.join(re.findall(r'\d+', tabela[3])))

                        lista_posicoes = []
                        for index, cell in enumerate(linhas_posicao_colunas_para_criar[linha_config]):
                            if cell is not None and cell == 'x':
                                lista_posicoes.append(['x', index, linha_largura_colunas[index], linha_formatacao_colunas[index]])
                            elif cell is not None:
                                lista_posicoes.append([cell, index, linha_largura_colunas[index]])
                        
                        # colocar cabeçalho
                        coluna = coluna_inicial
                        for i, posicao_config in enumerate(lista_posicoes):
                            nova_aba.cell(row=linha_inicial, column=coluna, value=str(cabecalho[lista_posicoes[i][1]]))
                            coluna+= 1
                        
                        # print('valor coluna inicial', coluna_inicial)
                        linha_inicial_temp = linha_inicial + 1
                        while True:
                            coluna_inicial_temp = coluna_inicial
                            try:
                                if linhas_tabela_principal[index_filhos_tabela][1] == 'x':
                                    index_filhos_tabela = index_principal
                                    break
                            except:
                                break
                            for i, posicao_config in enumerate(lista_posicoes):
                                valor = linhas_tabela_principal[index_filhos_tabela][lista_posicoes[i][1]]
                                if posicao_config[0] != 'x':
                                    valor = f"'{posicao_config[0]}"
                                cel = nova_aba.cell(row=linha_inicial_temp, column=coluna_inicial_temp, value=valor)
                                # colocando formato na celula
                                if not "=" in posicao_config[0] and posicao_config[3] is not None:
                                    cel.number_format = posicao_config[3]
                                    cel = nova_aba.cell(row=linha_inicial_temp+1, column=coluna_inicial_temp)
                                    cel.number_format = posicao_config[3]
                                # definindo largura coluna
                                nova_aba.column_dimensions[get_column_letter(coluna_inicial_temp)].width = posicao_config[2]
                                coluna_inicial_temp+= 1
                                coluna_backup = coluna_inicial_temp
                            index_filhos_tabela += 1
                            linha_inicial_temp += 1
                        
                        # colocando total row
                        coluna_inicial_temp = coluna_inicial
                        tem_total_row = False
                        for i, posicao_config in enumerate(lista_posicoes):
                            valor = f"'{linha_total_row[lista_posicoes[i][1]]}"
                                # Aplicando Total a linha total
                            if i == 0:
                                posicao_coluna_total_text = coluna_inicial_temp
                                posicao_linha_total_text = linha_inicial_temp
                            if posicao_config[0] == 'x' or "=" in posicao_config[0]:
                                if valor is not None and "=" in valor:
                                    tem_total_row = True
                                    nova_aba.cell(row=linha_inicial_temp, column=coluna_inicial_temp, value=valor)
                                coluna_inicial_temp+= 1

                        if tem_total_row:
                            # nova_aba.cell(row=linha_inicial_temp, column=coluna_inicial_temp, value='Total')
                            celula = nova_aba.cell(row=posicao_linha_total_text, column=posicao_coluna_total_text, value='Total')
                            celula.font = Font(bold=True)
                        else:
                            linha_inicial_temp -=1
                        print('Aplicando formatação...')
                        range_final = f'{tabela[3]}:{get_column_letter(coluna_backup-1)}{linha_inicial_temp-1}'
                        # print(f"{linha[4]}_{tabela[2]}")
                        # tab = Table(displayName=f"{linha[4]}_{tabela[2]}", ref=range_final, totalsRowCount=1, totalsRowShown=True)
                        tab = Table(displayName=f"{linha[4]}_{tabela[2]}", ref=range_final)
                        style = TableStyleInfo(
                            name=tabela[1]
                        )  # Alinhar o conteúdo ao topo)
                        tab.tableStyleInfo = style
                        tab.autofilter = False
                        nova_aba.add_table(tab)
                        for row in nova_aba.iter_rows(min_row=2, max_row=nova_aba.max_row, min_col=1, max_col=nova_aba.max_column):
                            for cell in row:
                                cell.alignment = cell.alignment.copy(vertical='top')

                        # for value in 
                    
                # break
            
        # Remover filtro de todas as tabelas
        print('Removendo filtros das tabelas...')
        for sheetname in wb.sheetnames:
            if sheetname != aba_principal:
                ws = wb[sheetname]
                for table in ws.tables.items():
                    tab = ws.tables[table[0]]
                    remove_table_filters(tab, ws)    
            else:
                continue            
        print('Salvando arquivo de saída...')
        # wb.save(arquivo_saida)
        
        # wb = load_workbook(arquivo_saida, data_only=True)
        ws = wb[aba_principal]

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            if sheetname != aba_principal:
                for table in ws.tables.values():
                    cabecalho_table = []
                    range_cabecalho = table.ref
                    # calculando linha de total
                    list_range = range_cabecalho.split(':')
                    coluna_range = re.findall(r'[a-zA-Z]', list_range[1])
                    coluna_range = "".join(coluna_range).strip()
                    linha_range = int(''.join(re.findall(r'\d+', list_range[1])))
                    range_cabecalho = f'{list_range[0]}:{coluna_range}{linha_range+1}'
                    for i, row in enumerate(ws[range_cabecalho]):
                        # lista de fórmulas

                        for f, cell in enumerate(row):
                            if "'=" in str(cell.value):
                                cell.value = transform_cell(cell.value, table.name)
                                # print(cell.value)
                    
                    table.ref = range_cabecalho
                    # totalsRowCount=1, totalsRowShown=True
                    # table.totalsRowCount = 1
                    # table.totalsRowShown = True
                    # table.showLastColumn=False
                    # table.showRowStripes=True
                    # table.showColumnStripes=True
                    # table.name = table.name + '1'
                    # ws.add_table(table)

                        
                                   
                #      break
                #         break
                # break
            
            # ws['AS29'] = '=SUBTOTAL(109,S1E1_tabelaY[FR])'
            # ws['AS29'] = '=SUBTOTAL(106,S1E1_tabelaY[FR])'
            # ws['AS29'] = '=SUBTOTAL(106,S1E1_tabelaY[qty;FR])'
            # ws.cell('AS29').value = '=SUBTOTAL(109,Table1[2011])'
                                
        
        wb.save(arquivo_saida)
        print('Arquivo criado com sucesso!')
        time.sleep(10)
    except:
        print(f'Erro ao executar o script, erro encontrado: {traceback.format_exc()}')
        time.sleep(120)