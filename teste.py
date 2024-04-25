from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.utils import get_column_letter

# Criar um novo arquivo do Excel e uma nova planilha
wb = Workbook()
ws = wb.active

# Exemplo de dados
data = [
    ['Produto', 'Quantidade', 'Preço'],
    ['Produto A', 10, 100],
    ['Produto B', 20, 200],
    ['Produto C', 30, 300]
]

# Preencher a planilha com os dados
for row in data:
    ws.append(row)

# Definir o intervalo da tabela
tabela = ws.dimensions

# Adicionar uma linha de total
coluna_total = 'C'  # Coluna para a qual você quer calcular o total
linha_total = len(data) + 1  # A linha após os dados

total_formula = f'=SUBTOTAL(109,{coluna_total}2:{coluna_total}{linha_total-1})'

# Definir a tabela
range_final = f'{tabela}'
tab = Table(displayName="Tabela1", ref=range_final)
tab.tableColumns = [
    TableColumn(id=idx+1, name=name) for idx, name in enumerate(data[0])
]

# Adicionar a fórmula de total para a coluna desejada
ws[f'{coluna_total}{linha_total}'] = total_formula

# Estilo da tabela
style = TableStyleInfo(name="TableStyleMedium9")
tab.tableStyleInfo = style

# Adicionar a tabela à planilha
ws.add_table(tab)

# Salvar o arquivo
wb.save("exemplo_tabela_com_total.xlsx")

