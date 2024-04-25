from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn, TableFormula
from openpyxl.worksheet.filters import AutoFilter

wb = Workbook()
ws = wb.active

# Define a formula to be used in a calculated column
formula_text = 'Table1[[#This Row], [2011]]+Table1[[#This Row], [2012]]+Table1[[#This Row], [2013]]+Table1[[#This Row], [2014]]'
ft = f'={formula_text}'

# Formulas needs to be added to manually to the table body. Formula must start with '=' here
data = [
    ['Apples', 10000, 5000, 8000, 6000, ft],
    ['Pears',   2000, 3000, 4000, 5000, ft],
    ['Bananas', 6000, 6000, 6500, 6000, ft],
    ['Oranges',  500,  300,  200,  700, ft],
    ['Total', None, None, None, None, None],
]

# add column headings. NB. these must be strings
ws.append(["Fruit", "2011", "2012", "2013", "2014", "Sum"])
for row in data:
    ws.append(row)

# Formula also needs to be added to manually to the column definition. Formula does not start with '=' here
data_formula = TableFormula(attr_text=formula_text)

'''
In order to configure totals row and add formulas we must define our own columns.
Column names used in formulas must match names defined here
'''
columns = [ 
    TableColumn(id=1, name='Fruit', totalsRowLabel='Total'),
    TableColumn(id=2, name='2011', totalsRowFunction='sum'),
    TableColumn(id=3, name='2012', totalsRowFunction='average'),
    TableColumn(id=4, name='2013'),
    TableColumn(id=5, name='2014'),
    TableColumn(id=6, name='Sum', calculatedColumnFormula=data_formula, totalsRowFunction='count'),
]

# Add subtotal function to totals row, refer to https://support.microsoft.com/en-gb/office/subtotal-function-7b027003-f060-4ade-9040-e478765b9939
ws.cell(row=6, column=2).value = '=SUBTOTAL(109,Table1[2011])'
ws.cell(row=6, column=3).value = '=SUBTOTAL(101,Table1[2012])'
ws.cell(row=6, column=6).value = '=SUBTOTAL(103,Table1[Sum])'

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)

# Create a table using custom columns, enable autofilter and totals row
auto_filter = AutoFilter(ref="A1:F6")
tab = Table(
    displayName="Table1", ref="A1:F6", tableStyleInfo = style, 
    totalsRowCount=1, totalsRowShown=True, 
    tableColumns=columns,
    autoFilter=auto_filter,
)
'''
Table must be added using ws.add_table() method to avoid duplicate names.
Using this method ensures table name is unque through out defined names and all other table name. 
'''
ws.add_table(tab)

wb.save("table2.xlsx")
